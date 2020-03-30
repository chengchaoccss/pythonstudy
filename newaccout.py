import openpyxl as xl
from openpyxl.styles import  PatternFill
import operator

def readcustom(filename):# 该函数来读取客户文件，提取每条金额以及对应的发票信息！
    wb = xl.load_workbook(f'{filename}.xlsx')
    sheet = wb['Sheet1']
    maxrow = sheet.max_row
    for row in range(2,maxrow+1):
        money = sheet[f'H{row}'].value
        invoice = sheet[f'D{row}'].value.split('/')
        print(f'{invoice} and {money}')
        # equalexist(row,sheet[f'D{row}'].value,money)
        # judgemoney(invoice,money)

def equalexist(row,value,money): #该函数为解析发票内容
    if '=' in value:
        inv = value.split('=')[-1].strip('\"').split(',')
        list= []
        newlist=[]
        print('+++++++++++++++++++++++++++++++++++++++++++++++')
        for i in inv:
            vo = i.split('-')[0]
            list.append(vo)
            newlist.append(vo)
            # print(vo)
        res.append(list)
        newlist.append(f'{money}')
        newres.append(newlist)
        # print('chang',len(res))
        summ=0
        newmoney.append(money)
        cout =0
        if row>=3:
            a= res[row-2]
            b = res[row-3]
            if operator.eq(a,b):
                cout+=1
                print(f'{row}行与{row-1}行相同！')
                global newrow
                newrow = row-1
                summ=newmoney[row-3]+newmoney[row-2]
                newmoney[row-2]=summ
                global newsum
                newsum = summ

            else:
                # print(a)
                if a==['\xa0 ']:
                    print(f'{row}行为空行!')
                    print(f'{newrow}~{row-1}的总和为{newsum}')
                else:
                    print(f'{row}行与{row-1}行不相同！')
                    print(f'{newrow}~{row-1}的总和为{newsum}')

    else:
        print(f'D{row}没有等号！')



def judgemoney(invoicelsit,money):# 将客户方的发票信息与自己的文件对照，判断金额是否正确！
    wb = xl.load_workbook('MSUself.xlsx')
    sheet = wb['应收款明细表']
    maxrow = sheet.max_row
    sum = 0
    fillnum = []
    for row in range(2,maxrow+1):
        if str(sheet[f'A{row}'].value) in invoicelsit:
            sum += sheet[f'B{row}'].value
            fillnum.append(row)

    if sum == money:
        # print(f'没有问题，该款明细正确！总额为{sum}')
        fillgreen(True,fillnum,money)
    else:
        # print(f'我方总额为：{sum}，客户方总额为:{money}')
        fillgreen(False,fillnum,money)

def fillgreen(type,fillnum,money): # 该函数是对正确的金额与错误的金额两种情况进行颜色填充，绿色表示金额正确，红色表示错误。
    wb = xl.load_workbook('MSUself.xlsx')
    sheet = wb['Sheet1']

    if type is True:
        fill = PatternFill("solid", fgColor="ADFF2F") # 颜色的值参考网址：http://www.114la.com/other/rgb.htm
        for i in fillnum:
            sheet[f'C{i}'].fill = fill
            sheet[f'C{i}'].value = money
        wb.save('MSUself.xlsx')

    else:
        fill = PatternFill("solid", fgColor="CD3333")
        for i in fillnum:
            sheet[f'C{i}'].fill = fill
            sheet[f'C{i}'].value = money
        wb.save('MSUself.xlsx')


def main():
    readcustom('MSU')

if __name__ == '__main__':
    res = []
    newres = []
    lastres = []
    newmoney =[]
    strlist=[]
    newstrlist=[]
    newsum = 0
    newrow =0
    main()

