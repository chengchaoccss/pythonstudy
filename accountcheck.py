import openpyxl as xl
from openpyxl.styles import  PatternFill
import operator

def readcustom(filename):# 该函数来读取客户文件，提取每条金额以及对应的发票信息！
    wb = xl.load_workbook(f'{filename}.xlsx')
    sheet = wb['Sheet1']
    maxrow = sheet.max_row
    for row in range(2,maxrow+1):
        money = sheet[f'H{row}'].value
        # invoice = sheet[f'D{row}'].value.split('/')
        equalexist(row,sheet[f'D{row}'].value,money)

def equalexist(row,value,money): #该函数为解析发票内容
    if '=' in value:
        inv = value.split('=')[-1].strip('\"').split(',')
        list= []
        newlist=[]
        for i in inv:
            vo = i.split('-')[0]
            list.append(vo)
            newlist.append(vo)
            # print(vo)
        res.append(list)
        newlist.append(f'{money}')
        newres.append(newlist)
        # print('chang',len(res))
        summ=0
        newmoney.append(money)
        cout =0
        if row>=3:
            a= res[row-2]
            b = res[row-3]
            if operator.eq(a,b):
                cout+=1
                # print(f'{row}行与{row-1}行相同！')
                global newrow
                newrow = row-1
                summ=newmoney[row-3]+newmoney[row-2]
                newmoney[row-2]=summ
                global newsum
                newsum = summ

            else:
                # print(a)
                if a==['\xa0 ']:
                    pass
                    # print(f'{row}行为空行!')
                    # print(f'{newrow}~{row-1}的总和为{newsum}')
                else:
                    pass
                    # print(f'{row}行与{row-1}行不相同！')
                    # print(f'{newrow}~{row-1}的总和为{newsum}')

    else:
        print(f'D{row}没有等号！')



def judgemoney(invoicelsit,money):# 将客户方的发票信息与自己的文件对照，判断金额是否正确！
    wb = xl.load_workbook('MSUself.xlsx')
    sheet = wb['应收款明细表']
    maxrow = sheet.max_row
    sum = 0
    fillnum = []
    for row in range(2,maxrow+1):
        if len(str(sheet[f'G{row}'].value))==7:
            newdata='0'+str(sheet[f'G{row}'].value)

        elif len(str(sheet[f'G{row}'].value))==6:
            newdata = '00' + str(sheet[f'G{row}'].value)
        else:
            newdata = str(sheet[f'G{row}'].value)

        if newdata in invoicelsit:
            sum += sheet[f'Q{row}'].value
            fillnum.append(row)

    sum = round(sum,2)
    if sum + money==0:
        print('right:',end=' ')
        print(f'行号为{inds},没有问题，该款明细正确！总额为{sum}')
        # print('没问题')
        fillgreen(True,fillnum,money)
        customfill(True,inds,money)
    else:
        for i in range(40):
            print(' ',end=' ')
        print('wrong:',end=' ')
        print(f'行号为{inds}金额不一致,我方总额为：{sum}，客户方总额为:{money}')
        if sum == 0:
            zerofill(inds,money)
        else:
            fillgreen(False,fillnum,money)
            customfill(False,inds,money)


def fillgreen(type,fillnum,money): # 该函数是对正确的金额与错误的金额两种情况进行颜色填充，绿色表示金额正确，红色表示错误。
    wb = xl.load_workbook('MSUself.xlsx')
    sheet = wb['应收款明细表']

    if type is True:
        fill = PatternFill("solid", fgColor="ADFF2F") # 颜色的值参考网址：http://www.114la.com/other/rgb.htm
        for i in fillnum:
            sheet[f'S{i}'].fill = fill
            sheet[f'S{i}'].value = money
        wb.save('MSUself.xlsx')

    else:
        fill = PatternFill("solid", fgColor="CD3333")
        for i in fillnum:
            sheet[f'S{i}'].fill = fill
            sheet[f'S{i}'].value = money
        wb.save('MSUself.xlsx')


def zerofill(inds,money):
    wb = xl.load_workbook('MSU.xlsx')
    sheet = wb['Sheet1']

    fill = PatternFill("solid", fgColor="FFFF00") # 颜色的值参考网址：http://www.114la.com/other/rgb.htm
    for i in inds:
        sheet[f'G{i}'].fill = fill
        sheet[f'G{i}'].value = money
    wb.save('MSU.xlsx')


def customfill(type,inds,money):

    wb = xl.load_workbook('MSU.xlsx')
    sheet = wb['Sheet1']

    if type is True:
        fill = PatternFill("solid", fgColor="ADFF2F") # 颜色的值参考网址：http://www.114la.com/other/rgb.htm
        for i in inds:
            sheet[f'G{i}'].fill = fill
            sheet[f'G{i}'].value = money
        wb.save('MSU.xlsx')

    else:
        fill = PatternFill("solid", fgColor="CD3333")
        for i in inds:
            sheet[f'G{i}'].fill = fill
            sheet[f'G{i}'].value = money
        wb.save('MSU.xlsx')


def main():
    readcustom('MSU')

    for i in res:
        lastres.append(' '.join(i).split()) # 对res去空白字符

    for i in lastres:
        if i not in newstrlist:
            newstrlist.append(i)  #newstrlist是去重以后的发票信息列表

    for i in newres:
        strlist.append(' '.join(i).split())  #strlist是包含发票与金额所有信息的列表

    for i in newstrlist:
        count = 0
        for j in strlist:
            if j[:-1:] == i:
                count+=round(float(j[len(j)-1]),2)
        importantdata.append((i,count))

    for row in range(0,len(importantdata)):
        global inds
        inds = [i + 2 for i, x in enumerate(lastres) if x == importantdata[row - 2][0]]  # 返回行数
        judgemoney(importantdata[row-2][0],round(importantdata[row-2][1],2))


if __name__ == '__main__':
    try:
        res = []
        newres = []
        lastres = []
        newmoney =[]
        strlist=[]
        newstrlist=[]
        newsum = 0
        newrow =0
        importantdata = []
        main()
    except:
        print('运行错误，请检查运行环境！')
