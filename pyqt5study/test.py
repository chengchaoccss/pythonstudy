from PyQt5.QtWidgets import *
import accountcheck
from PyQt5 import QtCore
import sys
import openpyxl as xl
from openpyxl.styles import  PatternFill
import operator




class Accountcheck(QWidget):
    def __init__(self):
        super().__init__()
        self.text = ''
        self.initUI()

    def initUI(self):
        self.label1 = QLabel('客户表格文件路径：')
        self.label2 = QLabel('我方表格文件路径：')
        self.label3 = QLabel('客户发票信息列号：')
        self.label4 = QLabel('客户金额信息列号：')
        self.label5 = QLabel('我方发票信息列号：')
        self.label6 = QLabel('我方金额信息列号：')

        self.line1 = QLineEdit()
        self.line2 = QLineEdit()
        self.line3 = QLineEdit()
        self.line4 = QLineEdit()
        self.line5 = QLineEdit()
        self.line6 = QLineEdit()

        self.button = QPushButton('开始对账')
        self.textline  =QTextEdit()

        self.layout = QGridLayout()
        self.layout.addWidget(self.label1,0,0)
        self.layout.addWidget(self.line1,0,1)
        self.layout.addWidget(self.label2,1,0)
        self.layout.addWidget(self.line2,1,1)
        self.layout.addWidget(self.label3,2,0)
        self.layout.addWidget(self.line3,2,1)
        self.layout.addWidget(self.label4,3,0)
        self.layout.addWidget(self.line4,3,1)
        self.layout.addWidget(self.label5,4,0)
        self.layout.addWidget(self.line5,4,1)
        self.layout.addWidget(self.label6,5,0)
        self.layout.addWidget(self.line6,5,1)
        self.layout.addWidget(self.button,6,0,1,2)
        self.layout.addWidget(self.textline,0,2,7,1)
        self.setLayout(self.layout)

    def dataget(self):
        filenamemy = self.line2.text()
        filenamecus = self.line1.text()
        # self.text += filename
        # self.text += '\n'
        # self.textline.setText(self.text)
        # print(filename)

        mylist = []  # 该列表存放我方表格中的发票信息
        mymoneylist = []  # 该列表存放我们金额信息，位置与mylist中的发票信息位置一一对应。
        wb = xl.load_workbook(filenamemy+'.xlsx')
        sheet = wb['应收款明细表']  # 注意这里的sheet名
        maxrow = sheet.max_row

        for row in range(2, maxrow + 1):
            money = sheet[f'Q{row}'].value
            invoice = str(sheet[f'G{row}'].value)
            if len(invoice) == 7:
                invoice = '0' + invoice
            elif len(invoice) == 6:
                invoice = '00' + invoice
            mylist.append(invoice)
            mymoneylist.append(money)
        print(mylist)
        # for i in mylist:
        #     self.textline.setText(i)
        # print(mylist)
        # self.textline.setText(str(mylist))
        wb = xl.load_workbook(filenamecus+'.xlsx', data_only=True)
        sheet = wb['Sheet1']
        maxrow = sheet.max_row
        cusmoney = []
        for i in range(2, maxrow + 1):
            cell = sheet[f'D{i}'].value
            if cell is not None:
                cmoney = round(float(sheet[f'D{i}'].value), 2)
                cusmoney.append(cmoney)
            else:
                cusmoney.append(cell)
        # return cusmoney
        print(cusmoney)

if __name__ == '__main__':
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)  # 解决了Qtdesigner设计的界面与实际运行界面不一致的问题
    app = QApplication(sys.argv)
    mainwin = Accountcheck()
    mainwin.button.clicked.connect(mainwin.dataget)
    mainwin.show()
    sys.exit(app.exec_())