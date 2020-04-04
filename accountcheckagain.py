import openpyxl as xl
from openpyxl.styles import  PatternFill
import operator

def myinvoice(): #该函数获取我方表格中的发票信息以及对应的金额信息。
    mylist = []  # 该列表存放我方表格中的发票信息
    mymoneylist = []  # 该列表存放我们金额信息，位置与mylist中的发票信息位置一一对应。
    wb = xl.load_workbook(f'{myfilename}.xlsx')
    sheet = wb['应收款明细表']  # 注意这里的sheet名
    maxrow = sheet.max_row

    for row in range(2,maxrow+1):
        money = sheet[f'Q{row}'].value
        invoice = str(sheet[f'G{row}'].value)
        if len(invoice) == 7:
            invoice='0'+invoice
        elif len(invoice) ==6:
            invoice='00'+invoice
        mylist.append(invoice)
        mymoneylist.append(money)
    return mylist,mymoneylist #两个列表对应位置分别是发票号以及金额

def customiovmoney(filename): #该函数处理客户表格，金额需要处于D列，会把这一列所有信息一次存储到cusmoney列表中。
    wb = xl.load_workbook(f'{filename}.xlsx',data_only=True)
    sheet = wb['Sheet1']
    maxrow = sheet.max_row
    cusmoney = []
    for i in range(2,maxrow+1):
        cell = sheet[f'D{i}'].value
        if cell is not None:
            cmoney = round(float(sheet[f'D{i}'].value),2)
            cusmoney.append(cmoney)
        else:
            cusmoney.append(cell)
    return cusmoney

def custominvinfo(filename): #该函数处理客户表格，发票信息处于B列，binfo是存储空白单元格的行号，invinfo是存储客户的发票号。
    wb = xl.load_workbook(f'{filename}.xlsx',data_only=True)
    sheet = wb['Sheet1']
    maxrow = sheet.max_row
    invinfo = []
    binfo= []   #binfo用来存储空白单元格的行号
    for i in range(2,maxrow+1):
        cell =sheet[f'B{i}'].value
        if len(str(cell)) == 7:
            cell = '0'+str(cell)
        elif len(str(cell))==6:
            cell='00'+str(cell)
        if cell is None:
            binfo.append(i)
            invinfo.append(cell)

        else:
            cell = str(cell).split('-')[0]
            invinfo.append(cell)
    return binfo,invinfo

def datadeal(binfo,cusmoney): #该函数是对客户表格中的每笔款项的总金额进行提取，存入countmoney，不是每笔发票的金额，而是多张发票的总金额
    countmoney = []
    for i in range(len(cusmoney)):
        if i + 2 in binfo:
            if cusmoney[i] is not None:
                countmoney.append(cusmoney[i])
    return countmoney


def checkaccount(custominvinfo,custommoneyinfo,myinvinfo,mymoneyinfo):#该函数较为重要，用来对账的。
    length = len(custommoneyinfo)
    teminvinfo = []#内部存储发票信息，但是每张发票都是唯一的，不存在重复的发票号。而且存储的金额也是每张发票的总金额，存储形式时【发票号 金额 发票号 金额】交替式。
    for i in range(length):
        if custominvinfo[i] is not None:
            if custominvinfo[i] in teminvinfo: # 发票号在teminvinfo中出现过就把金额数相加，赋值到旧的金额数上，这样就完成了同一张发票金额数统计的功能。
                ind = teminvinfo.index(custominvinfo[i])
                teminvinfo[ind+1]+=round(custommoneyinfo[i],2)
            else:
                teminvinfo.append(custominvinfo[i]) #如果这个发票不在teminvinfo中，表示搜索到目前为止，没有出现过，那么就把发票信息以及金额追究进去。
                teminvinfo.append(round(custommoneyinfo[i],2))

    for i in range(length):
        if custominvinfo[i] in myinvinfo:   #遍历客户的发票信息，判断是否在我方明细表中。
            ind =myinvinfo.index(custominvinfo[i]) #找到我方明细表中这张发票索引号
            ind2 = teminvinfo.index(custominvinfo[i]) #找到这张发票客户的金额数
            if mymoneyinfo[ind] == round(teminvinfo[ind2+1],2):
                print(f'没有问题，发票号为：{custominvinfo[i]},行号{ind+2}我方金额：{mymoneyinfo[ind]},对方金额：{round(teminvinfo[ind2 + 1],2)}')
                fillcell(True,i,mymoneyinfo[ind]) #在客户表格中进行相关填充标记！
                fillcellself(True,ind+2,round(teminvinfo[ind2+1],2))#在我方表格中对相关发票信息进行填充标记，填充的数字时客户方的金额数。
            else:
                fillcell(False, i, mymoneyinfo[ind])
                spacefill(40)

                fillcellself(False,ind+2,round(teminvinfo[ind2+1],2))
                print(f'有问题，发票号为：{custominvinfo[i]},      我方金额：{mymoneyinfo[ind]},对方金额：{round(teminvinfo[ind2+1],2)}')
        else:
            print(f'发票号{custominvinfo[i]}在我方明细表中没有找到！请核实发票信息！')

def spacefill(num): #用来打印空格的。
    for i in range(num):
        print(' ')

def fillcell(type,ind,money): # 该函数对客户表格进行标记，正确的金额与错误的金额两种情况进行颜色填充，绿色表示金额正确，红色表示错误。
    wb = xl.load_workbook(f'{customfilename}.xlsx')
    sheet = wb['Sheet1']

    if type is True:
        fill = PatternFill("solid", fgColor="ADFF2F") # 颜色的值参考网址：http://www.114la.com/other/rgb.htm
        sheet[f'E{ind+2}'].fill = fill
        sheet[f'E{ind+2}'].value = money
        wb.save(f'{customfilename}.xlsx')

    else:
        fill = PatternFill("solid", fgColor="CD3333")
        sheet[f'E{ind+2}'].fill = fill
        sheet[f'E{ind+2}'].value = money
        wb.save(f'{customfilename}.xlsx')


def fillcellself(type,ind,money): # 该函数是对我方表格进行标记，正确的金额与错误的金额两种情况进行颜色填充，绿色表示金额正确，红色表示错误。
    wb = xl.load_workbook(f'{myfilename}.xlsx')
    sheet = wb['应收款明细表']

    if type is True:
        fill = PatternFill("solid", fgColor="ADFF2F") # 颜色的值参考网址：http://www.114la.com/other/rgb.htm
        sheet[f'T{ind}'].fill = fill
        sheet[f'T{ind}'].value = money
        wb.save(f'{myfilename}.xlsx')

    else:
        fill = PatternFill("solid", fgColor="CD3333")
        sheet[f'T{ind}'].fill = fill
        sheet[f'T{ind}'].value = money
        wb.save(f'{myfilename}.xlsx')

def main():
    binfo,invinfo = custominvinfo(customfilename)
    cusmoney= customiovmoney(customfilename)
    datadeal(binfo,cusmoney)
    mylist,mymoneylist = myinvoice()  # mylist列表存放我方所有发票信息，mymoneylist存放对应的金额
    checkaccount(invinfo,cusmoney,mylist,mymoneylist)


if __name__ == '__main__':
    print('-----------------------------------------------------------------------------------------------------')
    print('注意事项如下：')
    print('''
         1. 我方明细表与对方明细表以及本程序必须防止在同一文件夹下；
         2. 输入表名时不需要输入后缀名；
         3. 我方表的sheet名为：应收款明细表，对方表的sheet名为：Sheet1；
         4. 我方表中发票信息在G列，金额信息在Q列。客户表中发票信息在B列，金额在D列，标记信息在E列。
         ''')
    print('-----------------------------------------------------------------------------------------------------')

    myfilename= input('请输入我方明细表名：')
    customfilename=input('请输入客户表名：')
    res = []
    lastres = []
    strlist=[]
    moneylist = []
    main()
